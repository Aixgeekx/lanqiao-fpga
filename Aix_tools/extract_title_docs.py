from pathlib import Path
import re

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TITLE = ROOT / "title"
OUT = ROOT / "Aix_tools" / "title_docs_extract.md"
KEYWORDS = [
    "FPGA", "CT137X", "Spartan", "RESET", "S1", "LED", "I2C", "UART",
    "DS1302", "ADC", "DAC", "EEPROM", "SRAM", "W25Q128", "数码管", "赛点",
    "提交", "比赛", "线上",
]


def clean(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"(证件号码\s*)\d[\dXx]{14,20}", r"\1[REDACTED_ID]", text)
    return text.strip()


def pages_to_scan(path: Path, page_count: int) -> tuple[list[int], str]:
    name = path.name.lower()
    if page_count <= 80 or "17444501" in name or "比赛手册" in name or "ct137x" in name or "seg_table" in name:
        return list(range(page_count)), "full"
    sample = sorted(set(list(range(min(5, page_count))) + [page_count - 1]))
    return sample, "sample"


def extract_pdf(path: Path, max_excerpt: int = 3000) -> str:
    lines = [f"## PDF: {path.relative_to(ROOT)}"]
    try:
        with pdfplumber.open(path) as pdf:
            page_count = len(pdf.pages)
            scan_pages, mode = pages_to_scan(path, page_count)
            all_text = []
            hits = []
            for page_index in scan_pages:
                page = pdf.pages[page_index]
                page_no = page_index + 1
                text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
                text = clean(text)
                all_text.append(text)
                lower = text.lower()
                for word in KEYWORDS:
                    if word.lower() in lower:
                        sample = next((ln for ln in text.splitlines() if word.lower() in ln.lower()), "")
                        hits.append((page_no, word, sample[:180]))
                        break
            merged = clean("\n\n".join(all_text))
            lines.append(f"- pages: {page_count}")
            lines.append(f"- scan_mode: {mode}")
            lines.append(f"- scanned_pages: {', '.join(str(i + 1) for i in scan_pages)}")
            lines.append(f"- extracted_chars: {len(merged)}")
            lines.append("- keyword_hits:")
            for page, word, sample in hits[:30]:
                lines.append(f"  - p{page} [{word}] {sample}")
            if not hits:
                lines.append("  - none")
            lines.append("- excerpt:")
            excerpt = merged[:max_excerpt] if merged else "[no extractable text]"
            lines.append("```text")
            lines.append(excerpt)
            lines.append("```")
    except Exception as exc:
        lines.append(f"- error: {type(exc).__name__}: {exc}")
    return "\n".join(lines)


def extract_xlsx(path: Path) -> str:
    lines = [f"## XLSX: {path.relative_to(ROOT)}"]
    try:
        wb = load_workbook(path, data_only=True)
        lines.append(f"- sheets: {', '.join(wb.sheetnames)}")
        for ws in wb.worksheets:
            lines.append(f"### Sheet: {ws.title} ({ws.max_row} x {ws.max_column})")
            lines.append("```text")
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                if any(values):
                    lines.append("\t".join(values))
            lines.append("```")
    except Exception as exc:
        lines.append(f"- error: {type(exc).__name__}: {exc}")
    return "\n".join(lines)


def main() -> None:
    parts = ["# title docs extraction", ""]
    for path in sorted(TITLE.rglob("*")):
        if path.suffix.lower() == ".pdf":
            parts.append(extract_pdf(path))
            parts.append("")
        elif path.suffix.lower() == ".xlsx":
            parts.append(extract_xlsx(path))
            parts.append("")
    OUT.write_text("\n".join(parts), encoding="utf-8")
    print(OUT)


if __name__ == "__main__":
    main()
